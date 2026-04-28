import json
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import simplekml
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..config import OUTPUT_DIR, load_config, safe_dumps

log = logging.getLogger(__name__)


FONT = "Calibri"
PRIMARY = "115E59"
ACCENT = "0D9488"
ACCENT_LIGHT = "CCFBF1"
GREEN = "059669"
AMBER = "D97706"
RED = "DC2626"
TEXT = "1F2937"
MUTED = "6B7280"
STRIPE = "F7F7F5"
WHITE = "FFFFFF"
WARM_GRAY = "F5F5F4"
BORDER_COLOR = "E5E5E3"
GROUP_HEADER_BG = "1A7A72"

HEADER_FONT = Font(name=FONT, bold=True, color=WHITE, size=10)
HEADER_FILL = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
HEADER_BORDER = Border(
    bottom=Side(style="thin", color=ACCENT),
    left=Side(style="hair", color="0E8A82"),
    right=Side(style="hair", color="0E8A82"),
)
GROUP_HEADER_FONT = Font(name=FONT, bold=True, color=WHITE, size=9)
GROUP_HEADER_FILL = PatternFill(start_color=GROUP_HEADER_BG, end_color=GROUP_HEADER_BG, fill_type="solid")
GROUP_HEADER_BORDER = Border(
    bottom=Side(style="thin", color=PRIMARY),
    left=Side(style="thin", color=PRIMARY),
    right=Side(style="thin", color=PRIMARY),
    top=Side(style="thin", color=PRIMARY),
)
BODY_FONT = Font(name=FONT, size=10, color=TEXT)
BODY_BORDER = Border(bottom=Side(style="hair", color=BORDER_COLOR))
STRIPE_FILL = PatternFill(start_color=STRIPE, end_color=STRIPE, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

TITLE_FONT = Font(name=FONT, bold=True, size=18, color=PRIMARY)
SUBTITLE_FONT = Font(name=FONT, size=11, color=MUTED)
SECTION_FONT = Font(name=FONT, bold=True, size=11, color=PRIMARY)
LABEL_FONT = Font(name=FONT, size=10, color=MUTED)
VALUE_FONT = Font(name=FONT, bold=True, size=10, color=TEXT)
STAT_VALUE_FONT = Font(name=FONT, bold=True, size=22, color=PRIMARY)
STAT_LABEL_FONT = Font(name=FONT, size=9, color=MUTED)

DAM_COLUMN_ORDER = [
    "id", "name", "status", "region", "lat", "lon", "elevation_m",
    "height_m", "capacity_mcm", "grid_distance_km", "year_built", "purpose",
]

DAM_COLUMN_LABELS = {
    "id": "ID",
    "name": "Dam Name",
    "status": "Status",
    "region": "Region",
    "lat": "Latitude",
    "lon": "Longitude",
    "elevation_m": "Elevation (m)",
    "height_m": "Height (m)",
    "capacity_mcm": "Capacity (MCM)",
    "surface_area_km2": "Area (km²)",
    "grid_distance_km": "Grid Distance (km)",
    "nearest_substation": "Nearest Substation",
    "year_built": "Year Built",
    "purpose": "Purpose",
    "river": "River",
    "basin": "Basin",
    "sources": "Sources",
    "alt_names": "Alt Names",
    "fill_rate_pct": "Fill Rate (%)",
    "hv_line_km": "HV Line (km)",
    "capacity_source": "Capacity Source",
    "review_category": "Review Category",
    "visual_verdict": "Visual Verdict",
    "visual_notes": "Visual Notes",
    "raw_name": "Raw Name",
    "elev_cop90_m": "Elev. COP90 (m)",
    "elev_delta_m": "Elev. Delta (m)",
    "elev_flag": "Elev. Flag",
    "source_count": "Source Count",
    "coord_spread_m": "Coord Spread (m)",
    "needs_review": "Needs Review",
    "substation_voltage_kv": "Substation Voltage (kV)",
}

DAM_COLUMN_WIDTHS = {
    "id": 12, "name": 32, "status": 16, "region": 28, "lat": 12, "lon": 12,
    "elevation_m": 14, "height_m": 12, "capacity_mcm": 15, "grid_distance_km": 17,
    "year_built": 12, "purpose": 18, "river": 20, "basin": 20, "sources": 18,
    "alt_names": 30, "visual_notes": 50, "visual_verdict": 28, "review_category": 18,
    "raw_name": 24, "nearest_substation": 30, "needs_review": 30,
}

NUMERIC_COLUMNS = {
    "lat", "lon", "elevation_m", "height_m", "capacity_mcm", "surface_area_km2",
    "grid_distance_km", "year_built", "fill_rate_pct", "hv_line_km",
    "elev_cop90_m", "elev_delta_m", "coord_spread_m", "source_count",
    "substation_voltage_kv",
}

HIDDEN_COLUMNS = {"alt_names", "satellite_view", "ceb_flag"}

DAM_COLUMN_GROUPS = [
    ("IDENTITY", {"id", "name", "status"}),
    ("LOCATION", {"region", "lat", "lon", "elevation_m"}),
    ("RESERVOIR", {"height_m", "capacity_mcm", "surface_area_km2", "fill_rate_pct"}),
    ("GRID", {"grid_distance_km", "hv_line_km", "nearest_substation", "substation_voltage_kv"}),
    ("METADATA", {"year_built", "purpose", "river", "basin", "sources", "alt_names", "raw_name", "capacity_source"}),
    ("DATA QUALITY", {"review_category", "visual_verdict", "visual_notes", "elev_cop90_m", "elev_delta_m", "elev_flag", "source_count", "coord_spread_m", "needs_review"}),
]


def _build_dam_col_defs(columns):
    ordered = [c for c in DAM_COLUMN_ORDER if c in columns and c not in HIDDEN_COLUMNS]
    rest = [c for c in columns if c not in ordered and c not in HIDDEN_COLUMNS]
    all_cols = ordered + rest

    defs = []
    for col in all_cols:
        label = DAM_COLUMN_LABELS.get(col, col.replace("_", " ").title())
        width = DAM_COLUMN_WIDTHS.get(col, max(len(label) + 4, 14))
        align = "right" if col in NUMERIC_COLUMNS else "left"
        if col in ("year_built", "status", "id"):
            align = "center"
        defs.append((col, label, width, align))
    return defs

PAIR_COLUMNS = [
    ("rank", "Rank", 8, "center"),
    ("upper_dam_name", "Upper Dam", 28, "left"),
    ("lower_dam_name", "Lower Dam", 34, "left"),
    ("head_m", "Head (m)", 12, "right"),
    ("distance_km", "Distance (km)", 14, "right"),
    ("distance_head_ratio", "Dist/Head Ratio", 15, "right"),
    ("energy_mwh_standard", "Energy (MWh)", 15, "right"),
    ("energy_mwh_variable_fill", "Energy - Variable Fill (MWh)", 24, "right"),
    ("upper_fill_rate_pct", "Upper Fill Rate (%)", 17, "right"),
    ("lower_fill_rate_pct", "Lower Fill Rate (%)", 17, "right"),
    ("composite_score", "Score", 10, "center"),
    ("data_quality", "Data Quality", 14, "center"),
    ("capex_per_mwh_usd", "CAPEX ($/MWh)", 17, "right"),
    ("capex_advantage_pct", "CAPEX Advantage", 15, "right"),
    ("lcoe_usd_per_mwh", "LCOE ($/MWh)", 16, "right"),
    ("tunneling_cost_usd", "Tunneling ($)", 17, "right"),
    ("substation_cost_usd", "Substation ($)", 17, "right"),
    ("grid_distance_km", "Grid Distance (km)", 17, "right"),
    ("upper_capacity_mcm", "Upper Capacity (MCM)", 19, "right"),
    ("lower_capacity_mcm", "Lower Capacity (MCM)", 19, "right"),
]

NUMBER_FORMATS = {
    "lat": "0.000000",
    "lon": "0.000000",
    "elevation_m": "#,##0",
    "height_m": "0.0",
    "capacity_mcm": "#,##0.0",
    "surface_area_km2": "0.0",
    "year_built": "0",
    "grid_distance_km": "0.0",
    "fill_rate_pct": "0.0",
    "hv_line_km": "0.0",
    "elev_cop90_m": "0.0",
    "elev_delta_m": "0.0",
    "coord_spread_m": "0.0",
    "head_m": "#,##0",
    "distance_km": "0.0",
    "distance_head_ratio": "0.0",
    "energy_mwh_standard": "#,##0",
    "energy_mwh_variable_fill": "#,##0",
    "upper_fill_rate_pct": "0.0",
    "lower_fill_rate_pct": "0.0",
    "composite_score": "0.000",
    "capex_per_mwh_usd": "$#,##0",
    "capex_advantage_pct": "0.0%",
    "lcoe_usd_per_mwh": "$#,##0",
    "tunneling_cost_usd": "$#,##0",
    "substation_cost_usd": "$#,##0",
    "upper_capacity_mcm": "#,##0",
    "lower_capacity_mcm": "#,##0",
    "rank": "0",
}


def generate_clean_outputs(dam_registry, scored_pairs, all_pairs=None, filtered_pairs=None):
    paths = {}
    paths["excel"] = _export_excel(dam_registry, scored_pairs, all_pairs, filtered_pairs)
    paths["json"] = _export_json(dam_registry, scored_pairs)
    paths["kml_3d"] = _export_3d_kml(dam_registry, scored_pairs)
    paths["geojson"] = _export_geojson(dam_registry, scored_pairs)
    log.info(f"Generated all outputs: {list(paths.keys())}")
    return paths


def _build_group_spans(col_defs):
    col_keys = [k for k, l, w, a in col_defs]
    group_membership = {}
    for group_name, group_cols in DAM_COLUMN_GROUPS:
        for col in group_cols:
            group_membership[col] = group_name

    spans = []
    current_group = None
    start_idx = None

    for i, key in enumerate(col_keys):
        group = group_membership.get(key, "OTHER")
        if group != current_group:
            if current_group is not None:
                spans.append((current_group, start_idx, i - 1))
            current_group = group
            start_idx = i
        if i == len(col_keys) - 1:
            spans.append((current_group, start_idx, i))

    return spans


def _add_grouped_header_row(ws, col_defs):
    ws.insert_rows(1)
    spans = _build_group_spans(col_defs)

    for group_name, start_idx, end_idx in spans:
        start_col = start_idx + 1
        end_col = end_idx + 1
        if start_col != end_col:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col,
            )
        cell = ws.cell(row=1, column=start_col, value=group_name)
        cell.font = GROUP_HEADER_FONT
        cell.fill = GROUP_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = GROUP_HEADER_BORDER
        for col_idx in range(start_col + 1, end_col + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = GROUP_HEADER_FILL
            c.border = GROUP_HEADER_BORDER

    ws.row_dimensions[1].height = 24


def _style_data_sheet(ws, col_defs, score_col=None, grouped_header=False):
    header_row = 2 if grouped_header else 1
    data_start_row = header_row + 1

    col_aligns = {i: align for i, (key, label, width, align) in enumerate(col_defs, 1)}

    for col_idx, cell in enumerate(ws[header_row], 1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
    ws.row_dimensions[header_row].height = 36

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row), data_start_row):
        fill = STRIPE_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, cell in enumerate(row, 1):
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BODY_BORDER
            h_align = col_aligns.get(col_idx, "left")
            cell.alignment = Alignment(horizontal=h_align, vertical="center")
        ws.row_dimensions[row_idx].height = 22

    for i, (key, label, width, align) in enumerate(col_defs, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width
        fmt = NUMBER_FORMATS.get(key)
        if fmt:
            for row_idx in range(data_start_row, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=i)
                if cell.value is not None:
                    if key == "capex_advantage_pct":
                        try:
                            cell.value = float(cell.value) / 100
                        except (ValueError, TypeError):
                            pass
                    cell.number_format = fmt

    if score_col:
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=score_col)
            if cell.value is None:
                continue
            try:
                v = float(cell.value)
            except (ValueError, TypeError):
                continue
            if v >= 0.7:
                cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
            elif v >= 0.5:
                cell.font = Font(name=FONT, size=10, bold=True, color=AMBER)
            else:
                cell.font = Font(name=FONT, size=10, bold=True, color=RED)

    dq_col = next((i for i, (k, l, w, a) in enumerate(col_defs, 1) if k == "data_quality"), None)
    if dq_col:
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=dq_col)
            v = str(cell.value).lower() if cell.value else ""
            if v == "complete":
                cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
            elif v == "partial":
                cell.font = Font(name=FONT, size=10, bold=True, color=AMBER)
            elif v == "missing":
                cell.font = Font(name=FONT, size=10, bold=True, color=RED)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = f"A{data_start_row}"


def _build_summary_sheet(wb, dam_registry, scored_pairs):
    ws = wb.create_sheet("Summary", 0)
    config = load_config()
    country = config.get("country", "Unknown")
    total_dams = len(dam_registry)
    total_pairs = total_dams * (total_dams - 1) // 2
    viable = len(scored_pairs)
    top_energy = scored_pairs["energy_mwh_standard"].max() if "energy_mwh_standard" in scored_pairs.columns and len(scored_pairs) > 0 else 0

    ws.sheet_properties.tabColor = ACCENT

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 18

    accent_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
    row = 1
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = accent_fill
    ws.row_dimensions[row].height = 4

    row = 3
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = ws.cell(row=row, column=2, value=f"PSH Screening Results: {country}")
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center")

    row = 4
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = ws.cell(row=row, column=2, value=f"Generated {datetime.now().strftime('%B %d, %Y')}")
    c.font = SUBTITLE_FONT

    row = 6
    stats = [
        (str(total_dams), "Dams Evaluated"),
        (f"{total_pairs:,}", "Pairs Generated"),
        (str(viable), "Viable Pairs"),
        (f"{top_energy:,.0f}", "Top Energy (MWh)"),
    ]
    stat_cols = [2, 3, 4, 5]

    stat_bg = PatternFill(start_color=ACCENT_LIGHT, end_color=ACCENT_LIGHT, fill_type="solid")
    stat_border = Border(
        top=Side(style="thin", color=ACCENT),
        bottom=Side(style="thin", color=ACCENT),
        left=Side(style="thin", color=ACCENT),
        right=Side(style="thin", color=ACCENT),
    )
    for i, (value, label) in enumerate(stats):
        col = stat_cols[i]
        c = ws.cell(row=row, column=col, value=value)
        c.font = STAT_VALUE_FONT
        c.fill = stat_bg
        c.border = stat_border
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=row + 1, column=col, value=label)
        c.font = STAT_LABEL_FONT
        c.fill = stat_bg
        c.border = stat_border
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 34
    ws.row_dimensions[row + 1].height = 18

    row = 9
    divider_fill = PatternFill(start_color=BORDER_COLOR, end_color=BORDER_COLOR, fill_type="solid")
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = divider_fill
    ws.row_dimensions[row].height = 2

    row = 11
    ws.cell(row=row, column=2, value="Filters & Physics").font = SECTION_FONT
    ws.cell(row=row, column=5, value="Cost Model").font = SECTION_FONT
    row += 1
    ph = config.get("physics", {})
    cm = config.get("cost_model", {})
    cb = config.get("cost_benchmarks", {})
    left_params = [
        ("Minimum head", f"{config['filters']['min_head_m']} m"),
        ("Maximum distance", f"{config['filters']['max_distance_km']} km"),
        ("Max dist/head ratio", f"{config['filters'].get('max_distance_head_ratio', 50)}"),
        ("Minimum capacity", f"{config['filters']['min_capacity_mcm']} MCM"),
        ("Round-trip efficiency", f"{int(ph.get('round_trip_efficiency', 0.78) * 100)}%"),
        ("Usable reservoir", f"{int(ph.get('usable_volume_fraction', 0.6) * 100)}%"),
        ("Power duration", f"{ph.get('power_duration_hours', 8)} hours"),
        ("Battery CAPEX", f"${cb.get('battery_capex_usd_per_mwh', 100_000):,.0f}/MWh"),
        ("Battery LCOE (ref)", f"${cb.get('battery_usd_per_mwh', 300)}/MWh"),
    ]
    right_params = [
        ("Penstock", f"${cm.get('penstock_usd_per_km', 46_960_000):,.0f}/km"),
        ("Upper reservoir", f"${cm.get('reservoir_upper_usd_per_mcm', 15_230_000):,.0f}/MCM"),
        ("Lower reservoir", f"${cm.get('reservoir_lower_usd_per_mcm', 38_770_000):,.0f}/MCM"),
        ("Powerhouse", f"${cm.get('powerhouse_usd_per_mw', 467_967):,.0f}/MW"),
        ("Substation", f"${cm.get('substation_usd_per_mw', 92_565):,.0f}/MW"),
        ("Roads (fixed)", f"${cm.get('roads_usd_fixed', 10_800_000):,.0f}"),
        ("Other", f"${cm.get('other_usd_per_mw', 66_852):,.0f}/MW"),
        ("Depreciation", f"{cm.get('depreciation_years', 40)} years"),
        ("Annual cycles", f"{cm.get('annual_cycles', 300)}"),
    ]
    for i, (label, value) in enumerate(left_params):
        ws.cell(row=row + i, column=2, value=label).font = LABEL_FONT
        ws.cell(row=row + i, column=3, value=value).font = VALUE_FONT
    for i, (label, value) in enumerate(right_params):
        ws.cell(row=row + i, column=5, value=label).font = LABEL_FONT
        ws.cell(row=row + i, column=6, value=value).font = VALUE_FONT

    row += max(len(left_params), len(right_params)) + 1
    ws.cell(row=row, column=2, value="Scoring Weights").font = SECTION_FONT
    row += 1
    for key, weight in config["scoring_weights"].items():
        label = key.replace("_", " ").title()
        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        c = ws.cell(row=row, column=3, value=f"{int(weight * 100)}%")
        c.font = VALUE_FONT
        row += 1

    row += 1
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = divider_fill
    ws.row_dimensions[row].height = 2

    row += 2
    ws.cell(row=row, column=2, value="Top 5 Pairs").font = SECTION_FONT
    row += 1

    top5_headers = ["Rank", "Upper Dam", "Lower Dam", "Head (m)", "Energy (MWh)", "Score"]
    top5_aligns = ["center", "left", "left", "right", "right", "center"]
    for i, header in enumerate(top5_headers):
        col = 2 + i
        c = ws.cell(row=row, column=col, value=header)
        c.font = Font(name=FONT, bold=True, size=9, color=WHITE)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = HEADER_BORDER
    ws.row_dimensions[row].height = 28
    row += 1

    for _, pair in scored_pairs.head(5).iterrows():
        values = [
            int(pair.get("rank") or 0),
            pair.get("upper_dam_name", ""),
            pair.get("lower_dam_name", ""),
            pair.get("head_m") or 0,
            pair.get("energy_mwh_standard") or 0,
            pair.get("composite_score") or 0,
        ]
        fmts = ["0", None, None, "#,##0", "#,##0", "0.000"]
        for i, (val, fmt) in enumerate(zip(values, fmts)):
            col = 2 + i
            c = ws.cell(row=row, column=col, value=val)
            c.font = BODY_FONT
            c.border = BODY_BORDER
            c.alignment = Alignment(horizontal=top5_aligns[i], vertical="center")
            if fmt:
                c.number_format = fmt
            if i == 5:
                try:
                    v = float(val)
                    if v >= 0.7:
                        c.font = Font(name=FONT, size=10, bold=True, color=GREEN)
                    elif v >= 0.5:
                        c.font = Font(name=FONT, size=10, bold=True, color=AMBER)
                    else:
                        c.font = Font(name=FONT, size=10, bold=True, color=RED)
                except (ValueError, TypeError):
                    pass
        if row % 2 == 0:
            for i in range(6):
                ws.cell(row=row, column=2 + i).fill = STRIPE_FILL
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = divider_fill
    ws.row_dimensions[row].height = 2

    row += 2
    ws.cell(row=row, column=2, value="Methodology").font = SECTION_FONT
    row += 1
    methodology = (
        "All dam pairs within the configured distance and head thresholds were evaluated. "
        "Each viable pair was scored on energy potential (40%), cost competitiveness vs battery "
        "storage CAPEX (35%), and grid proximity (25%). The composite score (0-1) ranks pairs "
        "by overall PSH viability."
    )
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=7)
    c = ws.cell(row=row, column=2, value=methodology)
    c.font = Font(name=FONT, size=9, color=MUTED)
    c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.sheet_view.showGridLines = False

    return ws


DH_COLUMNS = [
    ("dh_rank", "D/H Rank", 10, "center"),
    ("upper_dam_name", "Upper Dam", 28, "left"),
    ("lower_dam_name", "Lower Dam", 28, "left"),
    ("head_m", "Head (m)", 12, "right"),
    ("distance_km", "Distance (km)", 14, "right"),
    ("distance_head_ratio", "D/H Ratio", 12, "right"),
    ("upper_capacity_mcm", "Upper Cap (MCM)", 16, "right"),
    ("lower_capacity_mcm", "Lower Cap (MCM)", 16, "right"),
    ("grid_distance_km", "Grid Dist (km)", 14, "right"),
    ("screening_status", "Screening", 14, "center"),
    ("fail_reason", "Fail Reason", 52, "left"),
]

DH_NUMBER_FORMATS = {
    "head_m": "#,##0",
    "distance_km": "0.0",
    "distance_head_ratio": "0.00",
    "upper_capacity_mcm": "#,##0.0",
    "lower_capacity_mcm": "#,##0.0",
    "grid_distance_km": "0.0",
}


def _build_top20_dh_sheet(wb, all_pairs, scored_pairs, filtered_pairs=None):
    if all_pairs is None or len(all_pairs) == 0:
        return

    passed_keys = set()
    if scored_pairs is not None and len(scored_pairs) > 0:
        for _, r in scored_pairs.iterrows():
            passed_keys.add((r.get("upper_dam_id"), r.get("lower_dam_id")))

    filtered_reasons = {}
    fp = filtered_pairs
    if fp is None:
        from ..config import DATA_PROCESSED
        filtered_path = DATA_PROCESSED / "filtered_pairs.json"
        if filtered_path.exists():
            fp = pd.read_json(filtered_path)
    if fp is not None:
        for _, r in fp.iterrows():
            key = (r.get("upper_dam_id"), r.get("lower_dam_id"))
            reasons = r.get("tier1_reasons") or r.get("energy_filter_reasons") or ""
            if reasons and reasons != "all criteria met":
                filtered_reasons[key] = reasons

    df = all_pairs.copy()
    df = df[df["distance_head_ratio"].notna()].sort_values("distance_head_ratio").head(20).reset_index(drop=True)
    df["dh_rank"] = df.index + 1
    df["screening_status"] = df.apply(
        lambda r: "Passed" if (r.get("upper_dam_id"), r.get("lower_dam_id")) in passed_keys else "Failed", axis=1
    )
    df["fail_reason"] = df.apply(
        lambda r: filtered_reasons.get((r.get("upper_dam_id"), r.get("lower_dam_id")), ""), axis=1
    )

    available_keys = set(df.columns.tolist())
    col_defs = [(k, l, w, a) for k, l, w, a in DH_COLUMNS if k in available_keys]

    ws = wb.create_sheet("Top 20 by DH Ratio")
    ws.sheet_properties.tabColor = "8B4513"

    for col_idx, (key, label, width, align) in enumerate(col_defs, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 36

    for row_idx, (_, pair) in enumerate(df.iterrows(), 2):
        fill = STRIPE_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, (key, label, width, align) in enumerate(col_defs, 1):
            val = pair.get(key)
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BODY_BORDER
            cell.alignment = Alignment(horizontal=align, vertical="center")
            fmt = DH_NUMBER_FORMATS.get(key)
            if fmt and val is not None:
                cell.number_format = fmt
            if key == "screening_status":
                if val == "Passed":
                    cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
                elif val == "Failed":
                    cell.font = Font(name=FONT, size=10, bold=True, color=RED)
        ws.row_dimensions[row_idx].height = 22

    ws.auto_filter.ref = f"A1:{get_column_letter(len(col_defs))}21"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _export_excel(dam_registry, scored_pairs, all_pairs=None, filtered_pairs=None):
    output_path = OUTPUT_DIR / "results.xlsx"

    dam_col_defs = _build_dam_col_defs(dam_registry.columns.tolist())
    pair_col_defs = [(k, l, w, a) for k, l, w, a in PAIR_COLUMNS if k in scored_pairs.columns]

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        dam_df = dam_registry[[k for k, l, w, a in dam_col_defs]].copy()
        for col in dam_df.columns:
            if col in ("purpose", "alt_names", "sources"):
                dam_df[col] = dam_df[col].apply(lambda v: ", ".join(v) if isinstance(v, list) else v)
        dam_df.columns = [l for k, l, w, a in dam_col_defs]
        dam_df.to_excel(writer, sheet_name="Dam Registry", index=False)

        pair_df = scored_pairs[[k for k, l, w, a in pair_col_defs]].copy()
        for col in ["energy_mwh_standard"]:
            if col in pair_df.columns:
                pair_df[col] = pd.to_numeric(pair_df[col], errors="coerce").round(0)
        if "composite_score" in pair_df.columns:
            pair_df["composite_score"] = pair_df["composite_score"].round(3)
        pair_df.columns = [l for k, l, w, a in pair_col_defs]
        pair_df.to_excel(writer, sheet_name="Pairs Ranked", index=False)

        wb = writer.book

        _build_summary_sheet(wb, dam_registry, scored_pairs)
        _build_top20_dh_sheet(wb, all_pairs, scored_pairs, filtered_pairs)

        ws_dam = wb["Dam Registry"]
        _style_data_sheet(ws_dam, dam_col_defs)

        ws_pair = wb["Pairs Ranked"]
        score_col_idx = next((i for i, (k, l, w, a) in enumerate(pair_col_defs, 1) if k == "composite_score"), None)
        _style_data_sheet(ws_pair, pair_col_defs, score_col=score_col_idx)

    log.info(f"Saved Excel to {output_path}")
    return str(output_path)


def _export_json(dam_registry, scored_pairs):
    output_path = OUTPUT_DIR / "results.json"
    dam_list = []
    for _, row in dam_registry.iterrows():
        dam_list.append(row.to_dict())

    pairs_list = []
    for _, row in scored_pairs.iterrows():
        pairs_list.append(row.to_dict())

    result = {
        "generated": datetime.now().isoformat(),
        "config": load_config(),
        "total_dams": len(dam_registry),
        "total_pairs_scored": len(scored_pairs),
        "dams": dam_list,
        "pairs": pairs_list,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(safe_dumps(result, indent=2))
    log.info(f"Saved JSON to {output_path}")
    return str(output_path)


def _export_3d_kml(dam_registry, scored_pairs):
    dams = {d["name"]: d for _, d in dam_registry.iterrows()} if hasattr(dam_registry, "iterrows") else {}

    kml = simplekml.Kml(name="PSH Screening - Top Pairs 3D")

    for _, pair in scored_pairs.head(20).iterrows():
        pair = pair.to_dict()
        rank = int(pair.get("rank") or 0)
        upper_name = pair.get("upper_dam_name", "?")
        lower_name = pair.get("lower_dam_name", "?")
        head = pair.get("head_m") or 0
        dist = pair.get("distance_km") or 0
        energy = pair.get("energy_mwh_standard") or 0
        score = pair.get("composite_score") or 0

        folder = kml.newfolder(name=f"#{rank}: {upper_name} - {lower_name}")
        folder.description = f"Head: {head:.0f}m | Distance: {dist:.1f}km | Energy: {energy:,.0f} MWh | Score: {score:.3f}"

        upper_dam = dams.get(upper_name, {})
        lower_dam = dams.get(lower_name, {})
        upper_elev = pair.get("upper_elevation_m", 0)
        lower_elev = pair.get("lower_elevation_m", 0)

        for name, lat, lon, elev, dam_data, color, role in [
            (upper_name, pair["upper_lat"], pair["upper_lon"], upper_elev, upper_dam, simplekml.Color.red, "UPPER"),
            (lower_name, pair["lower_lat"], pair["lower_lon"], lower_elev, lower_dam, simplekml.Color.blue, "LOWER"),
        ]:
            pt = folder.newpoint(name=f"{role}: {name}")
            pt.coords = [(lon, lat)]
            pt.altitudemode = simplekml.AltitudeMode.clamptoground
            pt.style.iconstyle.color = color
            pt.style.iconstyle.scale = 1.5
            cap = dam_data.get("capacity_mcm", "?") if isinstance(dam_data, dict) else "?"
            grid = dam_data.get("grid_distance_km", "?") if isinstance(dam_data, dict) else "?"
            pt.description = f"Elevation: {elev:.0f}m\nCapacity: {cap} MCM\nGrid: {grid}km"

        line = folder.newlinestring(name=f"{head:.0f}m head, {dist:.1f}km")
        line.coords = [(pair["upper_lon"], pair["upper_lat"]), (pair["lower_lon"], pair["lower_lat"])]
        line.altitudemode = simplekml.AltitudeMode.clamptoground
        line.style.linestyle.width = 4
        if score > 0.7:
            line.style.linestyle.color = simplekml.Color.green
        elif score > 0.5:
            line.style.linestyle.color = simplekml.Color.orange
        else:
            line.style.linestyle.color = simplekml.Color.red

        center_lat = (pair["upper_lat"] + pair["lower_lat"]) / 2
        center_lon = (pair["upper_lon"] + pair["lower_lon"]) / 2
        lookat = simplekml.LookAt()
        lookat.latitude = center_lat
        lookat.longitude = center_lon
        lookat.range = dist * 1500
        lookat.tilt = 60
        folder.lookat = lookat

    kml_path = OUTPUT_DIR / "top_pairs_3d.kml"
    kml.save(str(kml_path))
    log.info(f"Saved 3D KML to {kml_path}")
    return str(kml_path)


def _export_geojson(dam_registry, scored_pairs):
    features = []
    for _, pair in scored_pairs.iterrows():
        features.append({
            "type": "Feature",
            "geometry": {
                "type": "LineString",
                "coordinates": [
                    [pair["upper_lon"], pair["upper_lat"]],
                    [pair["lower_lon"], pair["lower_lat"]],
                ],
            },
            "properties": {
                "rank": int(pair.get("rank") or 0),
                "upper_dam": pair.get("upper_dam_name"),
                "lower_dam": pair.get("lower_dam_name"),
                "head_m": pair.get("head_m"),
                "distance_km": pair.get("distance_km"),
                "distance_head_ratio": pair.get("distance_head_ratio"),
                "energy_mwh": pair.get("energy_mwh_standard"),
                "score": pair.get("composite_score"),
                "lcoe_usd_per_mwh": pair.get("lcoe_usd_per_mwh"),
            },
        })

    geojson = {"type": "FeatureCollection", "features": features}
    geojson_path = OUTPUT_DIR / "pairs.geojson"
    with open(geojson_path, "w", encoding="utf-8") as f:
        f.write(safe_dumps(geojson, indent=2))
    log.info(f"Saved GeoJSON to {geojson_path}")
    return str(geojson_path)


def generate_all_outputs(dam_registry, all_pairs_df, scored_pairs_df, sensitivity_results=None):
    return generate_clean_outputs(dam_registry, scored_pairs_df)
