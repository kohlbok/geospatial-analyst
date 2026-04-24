import logging

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from ..config import OUTPUT_DIR
from .export import (
    HEADER_FONT, HEADER_FILL, HEADER_BORDER,
    BODY_FONT, BODY_BORDER, STRIPE_FILL, WHITE_FILL,
    GREEN, AMBER, RED, MUTED,
)

log = logging.getLogger(__name__)

TIER1_DAM_COLUMNS = [
    ("dam_id", "Dam ID", 14),
    ("dam_name", "Dam Name", 40),
    ("dam_lat", "Latitude", 14),
    ("dam_lon", "Longitude", 14),
    ("viable_up", "Viable Up", 14),
    ("viable_down", "Viable Down", 15),
    ("max_head_up_m", "Max Head Up (m)", 19),
    ("best_up_dist_km", "Up Dist (km)", 16),
    ("max_head_down_m", "Max Head Down (m)", 21),
    ("best_down_dist_km", "Down Dist (km)", 18),
    ("kill_reason", "Kill Reason", 34),
]

CANDIDATE_COLUMNS = [
    ("_rank", "Rank", 10),
    ("composite_score", "Score", 12),
    ("dam_name", "Dam Name", 40),
    ("dam_id", "Dam ID", 14),
    ("direction", "Direction", 13),
    ("existing_dam_role", "Existing Role", 17),
    ("centroid_lat", "Candidate Lat", 16),
    ("centroid_lon", "Candidate Lon", 16),
    ("sink_elevation_m", "Sink Elev (m)", 16),
    ("saddle_elevation_m", "Saddle Elev (m)", 18),
    ("saddle_width_m", "Saddle Width (m)", 19),
    ("optimal_fill_m", "Optimal Fill Elev (m)", 24),
    ("head_m", "Head (m)", 14),
    ("distance_km", "Distance (km)", 17),
    ("distance_head_ratio", "D/H Ratio", 14),
    ("usable_volume_mcm", "Usable Volume (MCM)", 22),
    ("candidate_volume_mcm", "Candidate Vol (MCM)", 21),
    ("existing_capacity_mcm", "Existing Cap (MCM)", 20),
    ("binding_reservoir", "Binding Reservoir", 19),
    ("data_quality", "Data Quality", 16),
    ("grid_distance_km", "Grid Distance (km)", 21),
    ("penstock_length_m", "Penstock Length (m)", 24),
    ("optimal_diameter_m", "Penstock Diameter (m)", 24),
    ("flow_rate_m3s", "Flow Rate (m³/s)", 20),
    ("friction_pct", "Friction (%)", 16),
    ("net_head_m", "Net Head (m)", 16),
    ("energy_mwh", "Energy (MWh)", 18),
    ("energy_mwh_3hr", "Energy 3hr (MWh)", 20),
    ("energy_mwh_5hr", "Energy 5hr (MWh)", 20),
    ("energy_mwh_8hr", "Energy 8hr (MWh)", 20),
    ("max_discharge_hours", "Max Discharge (h)", 20),
    ("penstock_cost_usd", "Penstock Cost ($)", 20),
    ("reservoir_cost_usd", "Reservoir Cost ($)", 21),
    ("powerhouse_cost_usd", "Powerhouse Cost ($)", 22),
    ("substation_cost_usd", "Substation Cost ($)", 22),
    ("other_cost_usd", "Other Cost ($)", 18),
    ("total_capex_usd", "Total CapEx ($)", 20),
    ("capex_per_mwh_usd", "CapEx/MWh ($)", 18),
    ("battery_ratio", "Battery Ratio", 16),
    ("passes_battery_test", "Beats Battery", 16),
    ("abdelmoumen_flag", "Abdelmoumen Flag", 19),
    ("marginal_flag", "Marginal", 13),
    ("basin_capped", "Basin Capped", 14),
]


def generate_siting_excel(candidates, funnel_summary, dams_df=None, tier1_results=None, dam_scan_kills=None):
    wb = openpyxl.Workbook()

    _write_funnel_sheet(wb.active, funnel_summary)
    wb.active.title = "Funnel Summary"

    if dams_df is not None and tier1_results is not None:
        dam_funnel_sheet = wb.create_sheet("Dam Funnel")
        _write_dam_funnel_sheet(dam_funnel_sheet, dams_df, tier1_results, candidates, dam_scan_kills or {})

    candidates_sheet = wb.create_sheet("Candidates")
    _write_candidates_sheet(candidates_sheet, candidates)

    output_path = OUTPUT_DIR / "siting_results.xlsx"
    wb.save(str(output_path))
    log.info(f"Saved siting Excel to {output_path}")
    return output_path


def _t1_closest_attempt(t1):
    reason = t1.get("kill_reason")
    if reason == "ratio_too_high":
        parts = []
        h_up = t1.get("max_head_up_m")
        d_up = t1.get("best_up_dist_km")
        h_dn = t1.get("max_head_down_m")
        d_dn = t1.get("best_down_dist_km")
        if h_up and d_up:
            parts.append(f"D/H↑ {d_up * 1000 / h_up:.1f} — {d_up:.1f}km / {h_up:.0f}m")
        if h_dn and d_dn:
            parts.append(f"D/H↓ {d_dn * 1000 / h_dn:.1f} — {d_dn:.1f}km / {h_dn:.0f}m")
        return " | ".join(parts) if parts else "ratio too high"
    if reason == "flat_terrain":
        return "no 100m+ elevation change within 20km"
    if reason == "missing_coordinates_or_elevation":
        return "missing coordinates or elevation"
    if reason == "no_dem_data":
        return "no DEM tile available at Tier 1"
    return reason or ""


def _write_dam_funnel_sheet(ws, dams_df, tier1_results, candidates, dam_scan_kills):
    from openpyxl.styles import Font

    t1_by_id = {r["dam_id"]: r for r in tier1_results}
    candidate_dam_ids = {c["dam_id"] for c in candidates}
    best_capex_by_dam = {}
    for c in candidates:
        did = c["dam_id"]
        v = c.get("capex_per_mwh_usd")
        if v and (did not in best_capex_by_dam or v < best_capex_by_dam[did]):
            best_capex_by_dam[did] = v

    STAGE_COLORS = {
        "Candidate": GREEN,
        "T3": AMBER,
        "T2": RED,
        "T1": RED,
        "Excluded": MUTED,
    }

    cols = [
        ("Dam Name", 40),
        ("Funnel Stage", 26),
        ("Closest Attempt", 70),
        ("Best CapEx/MWh ($)", 22),
    ]
    for col_idx, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    for row_idx, (_, dam) in enumerate(dams_df.iterrows(), 2):
        dam_id = dam.get("id", "")
        name = dam.get("name", "")
        fill = STRIPE_FILL if row_idx % 2 == 0 else WHITE_FILL

        if dam_id in candidate_dam_ids:
            stage = "Candidate"
            detail = ""
            capex = best_capex_by_dam.get(dam_id)
        elif dam_id in dam_scan_kills:
            kill = dam_scan_kills[dam_id]
            stage = "T3" if "energy" in kill or "CapEx" in kill or "friction" in kill or "volume" in kill else "T2"
            detail = kill
            capex = None
        elif dam_id in t1_by_id:
            t1 = t1_by_id[dam_id]
            reason = t1.get("kill_reason")
            stage = "T1" if reason else "T2"
            detail = _t1_closest_attempt(t1) if reason else dam_scan_kills.get(dam_id, "scan not run")
            capex = None
        else:
            stage = "Excluded"
            detail = "existing PSH or missing coordinates"
            capex = None

        row = [name, stage, detail, capex]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BODY_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (2, 4) else "left",
                vertical="center",
                wrap_text=(col_idx == 3),
            )
            if col_idx == 2:
                color = STAGE_COLORS.get(stage[:2] if stage.startswith("T") else stage, MUTED)
                cell.font = Font(name="Calibri", size=10, color=color, bold=(stage == "Candidate"))
            if col_idx == 4 and isinstance(val, (int, float)):
                battery = 100_000
                color = GREEN if val < battery else (AMBER if val < battery * 1.5 else RED)
                cell.font = Font(name="Calibri", size=10, color=color, bold=(val < battery))


def _write_funnel_sheet(ws, funnel_summary):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 56

    headers = ["Stage", "Count", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("Total dams in registry", funnel_summary.get("total_dams", 0), ""),
        ("Phase 1 paired (map overlay)", funnel_summary.get("paired_excluded", 0), "Already have an existing-pair option; still scanned for greenfield"),
        ("Dams screened (Tier 1 input)", funnel_summary.get("tier1_input", 0), "Excludes existing PSH dams only"),
        ("Tier 1 survivors", funnel_summary.get("tier1_survivors", 0), "Have viable head in at least one direction"),
        ("Tier 1 killed: flat terrain", funnel_summary.get("killed_flat", 0), "No 100m+ elevation change within 20km"),
        ("Tier 1 killed: ratio too high", funnel_summary.get("killed_ratio", 0), "Elevation change only at uneconomic distance"),
        ("Tier 1 killed: other", funnel_summary.get("killed_other_t1", 0), "Missing data or other"),
        ("Dams with candidates", funnel_summary.get("dams_with_candidates", 0), "Produced >=1 viable greenfield candidate"),
        ("Viable candidates", funnel_summary.get("viable_candidates", 0), "Pass battery CapEx test or within 1.5x"),
        ("Beats battery benchmark", funnel_summary.get("beats_battery", 0), "CapEx/MWh < $100k"),
        ("Marginal (within 1.5x battery)", funnel_summary.get("marginal", 0), "Between $100k and $150k/MWh"),
    ]

    for row_idx, (label, count, notes) in enumerate(rows, 2):
        fill = STRIPE_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col, val in enumerate([label, count, notes], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BODY_BORDER
            if col == 2:
                cell.alignment = Alignment(horizontal="center")

    kill_reasons = funnel_summary.get("kill_reasons", {})
    if kill_reasons:
        start_row = len(rows) + 4
        ws.cell(row=start_row, column=1, value="Optimizer Kill Reasons").font = HEADER_FONT
        ws.cell(row=start_row, column=1).fill = HEADER_FILL
        ws.cell(row=start_row, column=2, value="Count").font = HEADER_FONT
        ws.cell(row=start_row, column=2).fill = HEADER_FILL
        for i, (reason, count) in enumerate(kill_reasons.items(), 1):
            ws.cell(row=start_row + i, column=1, value=reason).font = BODY_FONT
            ws.cell(row=start_row + i, column=2, value=count).font = BODY_FONT


def _write_candidates_sheet(ws, candidates):
    from openpyxl.styles import Font

    ws.row_dimensions[1].height = 30

    for col_idx, (field, label, width) in enumerate(CANDIDATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    sorted_candidates = sorted(
        candidates,
        key=lambda c: -(c.get("composite_score") or 0),
    )

    battery_capex = 100_000

    for row_idx, (rank, cand) in enumerate(enumerate(sorted_candidates, 1), 2):
        fill = STRIPE_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, (field, label, _) in enumerate(CANDIDATE_COLUMNS, 1):
            val = rank if field == "_rank" else cand.get(field)

            if isinstance(val, bool):
                display = "Yes" if val else "No"
            elif val is None:
                display = ""
            elif isinstance(val, float):
                display = round(val, 3)
            else:
                display = val

            cell = ws.cell(row=row_idx, column=col_idx, value=display)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BODY_BORDER

            if field == "capex_per_mwh_usd" and isinstance(val, (int, float)):
                if val < battery_capex:
                    cell.font = Font(name="Calibri", size=10, color=GREEN, bold=True)
                elif val < battery_capex * 1.5:
                    cell.font = Font(name="Calibri", size=10, color=AMBER, bold=True)
                else:
                    cell.font = Font(name="Calibri", size=10, color=RED)

            if field == "passes_battery_test":
                if val is True:
                    cell.font = Font(name="Calibri", size=10, color=GREEN, bold=True)
                elif val is False:
                    cell.font = Font(name="Calibri", size=10, color=RED)

            if field == "data_quality":
                if val == "complete":
                    cell.font = Font(name="Calibri", size=10, color=GREEN)
                elif val == "partial":
                    cell.font = Font(name="Calibri", size=10, color=AMBER, bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CANDIDATE_COLUMNS))}{len(sorted_candidates) + 1}"


def generate_tier1_excel(tier1_results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tier 1 Funnel"

    survivors = [r for r in tier1_results if r.get("viable_up") or r.get("viable_down")]
    killed = [r for r in tier1_results if not r.get("viable_up") and not r.get("viable_down")]
    kill_counts = {}
    for r in killed:
        reason = r.get("kill_reason", "other") or "other"
        kill_counts[reason] = kill_counts.get(reason, 0) + 1

    from openpyxl.styles import Font

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 56

    for col, h in enumerate(["Stage", "Count", "Notes"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")

    funnel_rows = [
        ("Total dams screened", len(tier1_results), ""),
        ("Tier 1 survivors", len(survivors), "Have 100m+ head within ratio limit"),
        ("Killed: flat terrain", kill_counts.get("flat_terrain", 0), "No 100m+ elevation change within 20km"),
        ("Killed: ratio too high", kill_counts.get("ratio_too_high", 0), "Elevation only reachable at uneconomic distance"),
        ("Killed: missing data", kill_counts.get("missing_coordinates_or_elevation", 0), "No coordinates or elevation"),
        ("Killed: no DEM data", kill_counts.get("no_dem_data", 0), "SRTM tile unavailable"),
        ("Killed: other", kill_counts.get("other", 0), ""),
    ]

    for row_idx, (label, count, notes) in enumerate(funnel_rows, 2):
        fill = STRIPE_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col, val in enumerate([label, count, notes], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BODY_BORDER
            if col == 2:
                cell.alignment = Alignment(horizontal="center")

    dam_sheet = wb.create_sheet("Dam Results")
    dam_sheet.row_dimensions[1].height = 30

    for col_idx, (field, label, width) in enumerate(TIER1_DAM_COLUMNS, 1):
        cell = dam_sheet.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        dam_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    sorted_results = sorted(tier1_results, key=lambda r: (
        0 if (r.get("viable_up") or r.get("viable_down")) else 1,
        -(r.get("max_head_up_m") or 0) - (r.get("max_head_down_m") or 0),
    ))

    for row_idx, record in enumerate(sorted_results, 2):
        fill = STRIPE_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, (field, label, _) in enumerate(TIER1_DAM_COLUMNS, 1):
            val = record.get(field)
            if isinstance(val, bool):
                display = "Yes" if val else "No"
            elif val is None:
                display = ""
            elif isinstance(val, float):
                display = round(val, 2)
            else:
                display = val
            cell = dam_sheet.cell(row=row_idx, column=col_idx, value=display)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BODY_BORDER
            if field in ("viable_up", "viable_down") and isinstance(val, bool):
                cell.font = Font(name="Calibri", size=10, color=GREEN if val else MUTED)

    dam_sheet.freeze_panes = "A2"
    dam_sheet.auto_filter.ref = f"A1:{get_column_letter(len(TIER1_DAM_COLUMNS))}{len(sorted_results) + 1}"

    output_path = OUTPUT_DIR / "siting_tier1.xlsx"
    wb.save(str(output_path))
    log.info(f"Saved Tier 1 Excel to {output_path}")
    return output_path


